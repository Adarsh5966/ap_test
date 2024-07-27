import openpyxl


wb = openpyxl.load_workbook('test_data.xlsx')
sheetnames = wb.sheetnames
print(sheetnames)

sh1 = wb['Sheet1']
for i in range(2,sh1.max_row+1):
    payload = {
        sh1.cell(row=1,column=1).value : sh1.cell(row=i,column=1).value
    }
print(payload)