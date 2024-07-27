import openpyxl


def excel_to_dict(file_path, sheet_name):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)
    # Select the desired sheet
    sheet = workbook[sheet_name]

    # Initialize an empty dictionary to store the values
    data_dict = {}

    max_row = sheet.max_row

    i=0
    # Iterate through each row in the sheet
    while i<max_row:
        for column in sheet.iter_cols(values_only=True):
            # Assuming the first column contains keys and the second column contains values

                key = column[0]  # First column
                value = column[1]  # Second column
                data_dict[key] = value
    return data_dict


# Example usage
file_path = "test_data.xlsx"
sheet_name = "Sheet1"
data = excel_to_dict(file_path, sheet_name)
print(data)
